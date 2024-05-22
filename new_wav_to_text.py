import openpyxl

def reverse_word(word):
    return word[::-1]

def text_to_excel(text_file_path, excel_file_path):
    # יצירת קובץ אקסל חדש או פתיחת קובץ קיים
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

    # קריאת קובץ הטקסט
    with open(text_file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    # העתקת השורות מקובץ הטקסט לקובץ האקסל
    row_num = 2  # התחלה משורה 2
    for line in lines:
        words = line.strip().split()  # פיצול השורה למילים
        reversed_words = [reverse_word(word) for word in words]  # הפיכת כל מילה
        reversed_line = ' '.join(reversed_words)  # איחוד המילים חזרה לשורה
        sheet.cell(row=row_num, column=1).value = reversed_line
        row_num += 4  # כל מעבר שורה בקובץ טקסט יתקדם 4 שורות באקסל

    # שמירת קובץ האקסל
    workbook.save(excel_file_path)
    print(f"Data has been written to {excel_file_path}")

# דוגמה לשימוש בפונקציה
#text_to_excel("new_text_song.txt", "chords.xlsx")
